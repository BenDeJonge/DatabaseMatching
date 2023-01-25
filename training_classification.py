# -*- coding: utf-8 -*-
"""
Transpose a table of trainings and followers.

Given are several .xslx files containing:
    n trainings with required followers.
These tables should be transposed to:
    a list of m jobtypes with required trainings.

    T1  T2  ... Tn
J1  T   V   ... T
J2  V   N   ... N
... ... ... ... ...
Jm  T   T   ... T

With the following abbreviations:
    Ji : the i-th job
    Ti : the i-th training
    T  : to be trained on x/y parts of the training
    V  : to be viewed on x/y parts of the training
    N  : not to be trained
"""

#-- I M P O R T S -------------------------------------------------------------
import openpyxl as xl
from re import split

#-- I N P U T S ---------------------------------------------------------------
xlsx_paths = ('process_master_sheet_1.xlsx', 'process_master_sheet_2.xlsx')
training_col = 'A'
follower_col = 'B'
viewing_col = 'C'
start_row = 2

output_file_name = 'test.xlsx'

#-- C O N S T A N T S ---------------------------------------------------------
# Tracking all existing jobs and trainings.
all_training_types = ('train', 'watch')
jobs : dict = {}
all_trainings : dict = {}

# Loading in workbook.
for xlsx_path in xlsx_paths:
    wb = xl.load_workbook(xlsx_path)
    ws = wb.active
    # Grabbing relevant columns.
    xlsx_follower_col = [ cell.value for cell in ws[follower_col][start_row-1:] ]
    xlsx_training_col = [ cell.value for cell in ws[training_col][start_row-1:] ]
    xlsx_viewing_col  = [ cell.value for cell in ws[viewing_col][start_row-1:] ]
    
    #-- M A I N   L O O P -----------------------------------------------------
    # Looping over all columns.
    for training, trainees, watchers in zip(xlsx_training_col,
                                            xlsx_follower_col,
                                            xlsx_viewing_col):
        # Increment occurences of training.
        try:
            all_trainings[training] += 1
        except KeyError:
            all_trainings[training] = 1
        # Find and format all followers of a training.
        for followers, training_type in zip((trainees, watchers), 
                                            all_training_types):
            # Training without any participants.
            if not followers:
                continue
            # Training with participants.
            followers = split(',|\n', followers)
            followers = [ f.strip() for f in followers ]
            for follower in followers:
                # The follower is not yet known.
                if not follower in jobs:
                    jobs[follower] = {training_type : {}
                                      for training_type in all_training_types}
                # Increment the training occurence of this type inside the job.
                try:
                    jobs[follower][training_type][training] += 1
                # The first time this training of this type is found for this job.
                except KeyError:
                    jobs[follower][training_type][training] = 1
    wb.close()
                
#-- W R I T I N G   O U T P U T -----------------------------------------------
# Creating output workbook.
wb = xl.Workbook()
ws = wb.active

# Writing headers in alphabetical order.
JOBCOL = 1
all_trainings = dict(sorted(all_trainings.items()))
ws['A1'] = 'Job'
for col, training in enumerate(all_trainings):
    _ = ws.cell(row=1, column=JOBCOL + col + 1, value = training)

# Writing data.
for i, (job, training_overview) in enumerate(jobs.items()):
    # Writing the job id.
    row = i + 2
    _ = ws.cell(row=row, column=JOBCOL, value=job)
    for training_type in all_training_types:
        # Getting all trainings for that job.
        trainings = training_overview[training_type]
        for training, amount in trainings.items():       
            # Writing the training type for each followed or watched training.
            column = list(all_trainings.keys()).index(training) + 1 + JOBCOL
            if not ws.cell(row=row, column=column).value:
                _ = ws.cell(row=row, 
                            column=column,
                            value=f'{training_type} {amount}/{all_trainings[training]}')       
# Outputting the file.
wb.save(output_file_name)